import io
import pandas as pd
from datetime import datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from typing import Optional


def generate_pdf_report(
    trades: list[dict],
    summary: dict,
    period: str,
    user_email: str,
) -> bytes:
    """Generate a professional PDF trading report."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=0.5*inch, bottomMargin=0.5*inch)
    styles = getSampleStyleSheet()
    story = []

    # ─── Header ───
    title_style = ParagraphStyle("title", parent=styles["Title"], fontSize=22, textColor=colors.HexColor("#0f172a"))
    story.append(Paragraph("TradeMinds — Trading Report", title_style))
    story.append(Paragraph(f"Period: {period} | Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", styles["Normal"]))
    story.append(Spacer(1, 0.3*inch))

    # ─── Summary Cards ───
    summary_data = [
        ["Metric", "Value"],
        ["Total Trades", str(summary.get("total_trades", 0))],
        ["Winning Trades", str(summary.get("winning_trades", 0))],
        ["Losing Trades", str(summary.get("losing_trades", 0))],
        ["Win Rate", f"{summary.get('win_rate', 0):.1f}%"],
        ["Total P&L", f"{summary.get('total_pnl', 0):+.2f} {summary.get('currency', 'USD')}"],
        ["Best Trade", f"{summary.get('best_trade', 0):+.2f}"],
        ["Worst Trade", f"{summary.get('worst_trade', 0):+.2f}"],
        ["Max Drawdown", f"{summary.get('max_drawdown', 0):.2f}%"],
        ["Profit Factor", f"{summary.get('profit_factor', 0):.2f}"],
        ["Sharpe Ratio", f"{summary.get('sharpe_ratio', 0):.2f}"],
        ["Avg Trade Duration", f"{summary.get('avg_duration_hours', 0):.1f}h"],
    ]

    summary_table = Table(summary_data, colWidths=[3*inch, 2.5*inch])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 11),
        ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#f8fafc")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f5f9")]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e2e8f0")),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
        ("FONTSIZE", (0, 1), (-1, -1), 10),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 0.3*inch))

    # ─── Trade History Table ───
    story.append(Paragraph("Trade History", styles["Heading2"]))
    story.append(Spacer(1, 0.1*inch))

    if trades:
        trade_headers = ["Date", "Symbol", "Side", "Entry", "Exit", "Lot", "P&L", "Strategy"]
        trade_rows = [trade_headers]
        for t in trades[:100]:  # Max 100 rows in PDF
            pnl = t.get("pnl", 0) or 0
            pnl_str = f"{pnl:+.2f}"
            trade_rows.append([
                t.get("opened_at", "")[:10],
                t.get("symbol", ""),
                t.get("side", "").upper(),
                str(t.get("entry_price", "")),
                str(t.get("exit_price", "") or "Open"),
                str(t.get("lot_size", "")),
                pnl_str,
                t.get("strategy_name", "")[:15],
            ])

        trade_table = Table(trade_rows, colWidths=[0.9*inch]*8)
        trade_style = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e293b")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.3, colors.HexColor("#e2e8f0")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
            ("ALIGN", (3, 1), (-1, -1), "RIGHT"),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]
        # Color P&L column
        for i, t in enumerate(trades[:100], start=1):
            pnl = t.get("pnl", 0) or 0
            color = colors.HexColor("#16a34a") if pnl >= 0 else colors.HexColor("#dc2626")
            trade_style.append(("TEXTCOLOR", (6, i), (6, i), color))

        trade_table.setStyle(TableStyle(trade_style))
        story.append(trade_table)

    doc.build(story)
    return buf.getvalue()


def generate_excel_report(
    trades: list[dict],
    summary: dict,
    period: str,
) -> bytes:
    """Generate a detailed Excel report with charts."""
    wb = Workbook()

    # ─── Summary Sheet ───
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.column_dimensions["A"].width = 25
    ws_summary.column_dimensions["B"].width = 20

    header_fill = PatternFill("solid", fgColor="0F172A")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    value_font = Font(size=11)
    center = Alignment(horizontal="center")

    ws_summary["A1"] = "TradeMinds — Trading Report"
    ws_summary["A1"].font = Font(bold=True, size=16, color="0F172A")
    ws_summary["A2"] = f"Period: {period}"
    ws_summary["A3"] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"
    ws_summary.append([])

    metrics = [
        ("Total Trades", summary.get("total_trades", 0)),
        ("Winning Trades", summary.get("winning_trades", 0)),
        ("Losing Trades", summary.get("losing_trades", 0)),
        ("Win Rate (%)", f"{summary.get('win_rate', 0):.1f}"),
        ("Total P&L", f"{summary.get('total_pnl', 0):+.2f}"),
        ("Best Trade", f"{summary.get('best_trade', 0):+.2f}"),
        ("Worst Trade", f"{summary.get('worst_trade', 0):+.2f}"),
        ("Max Drawdown (%)", f"{summary.get('max_drawdown', 0):.2f}"),
        ("Profit Factor", f"{summary.get('profit_factor', 0):.2f}"),
        ("Sharpe Ratio", f"{summary.get('sharpe_ratio', 0):.2f}"),
        ("Avg Duration (h)", f"{summary.get('avg_duration_hours', 0):.1f}"),
    ]

    for i, (metric, value) in enumerate(metrics, start=5):
        ws_summary.cell(row=i, column=1, value=metric).font = Font(bold=True)
        ws_summary.cell(row=i, column=2, value=value).font = value_font

    # ─── Trade Log Sheet ───
    ws_trades = wb.create_sheet("Trade Log")
    headers = ["Date", "Symbol", "Market", "Side", "Entry Price", "Exit Price",
               "Lot Size", "P&L", "P&L %", "Stop Loss", "Take Profit",
               "Strategy", "AI Confidence", "Closed By", "Duration (h)"]

    for col, header in enumerate(headers, 1):
        cell = ws_trades.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        ws_trades.column_dimensions[cell.column_letter].width = 14

    for row, trade in enumerate(trades, start=2):
        pnl = trade.get("pnl", 0) or 0
        values = [
            trade.get("opened_at", "")[:19],
            trade.get("symbol", ""),
            trade.get("market_type", ""),
            trade.get("side", "").upper(),
            trade.get("entry_price", ""),
            trade.get("exit_price", ""),
            trade.get("lot_size", ""),
            pnl,
            trade.get("pnl_pct", ""),
            trade.get("stop_loss", ""),
            trade.get("take_profit", ""),
            trade.get("strategy_name", ""),
            trade.get("ai_confidence", ""),
            trade.get("closed_by", ""),
            trade.get("duration_hours", ""),
        ]
        for col, value in enumerate(values, 1):
            cell = ws_trades.cell(row=row, column=col, value=value)
            if col == 8 and isinstance(pnl, (int, float)):
                cell.font = Font(
                    color="16A34A" if pnl >= 0 else "DC2626",
                    bold=True
                )

    # ─── Equity Curve Sheet ───
    ws_equity = wb.create_sheet("Equity Curve")
    ws_equity["A1"] = "Date"
    ws_equity["B1"] = "Equity"
    ws_equity["C1"] = "Cumulative P&L"

    cumulative = 0
    for i, trade in enumerate(trades, start=2):
        pnl = trade.get("pnl", 0) or 0
        cumulative += pnl
        ws_equity.cell(row=i, column=1, value=trade.get("closed_at", "")[:10])
        ws_equity.cell(row=i, column=2, value=cumulative + summary.get("initial_balance", 10000))
        ws_equity.cell(row=i, column=3, value=cumulative)

    if len(trades) > 1:
        chart = LineChart()
        chart.title = "Equity Curve"
        chart.style = 10
        chart.y_axis.title = "Balance"
        chart.x_axis.title = "Trade"
        data = Reference(ws_equity, min_col=2, min_row=1, max_row=len(trades)+1)
        chart.add_data(data, titles_from_data=True)
        ws_equity.add_chart(chart, "E5")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
